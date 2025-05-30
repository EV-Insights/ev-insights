import math
import os
import json
import textwrap
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
from pprint import pprint
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.text import LineBreak
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from zipfile import ZipFile
from io import BytesIO
from typing import List, Tuple, Union
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.platypus import PageBreak

from src.analysis.analysis import Analysis


class stats(Analysis):
    def __init__(self, id, name, info, enabled, full_custom_mode, show_images, save_images, save_results, input_interface, output_interface,
                 output_dir, data_selection, custom_params):
        super().__init__(id=id, name=name, info=info, enabled=enabled, full_custom_mode=full_custom_mode, show_images=show_images,
                         save_images=save_images, save_results=save_results, input_interface=input_interface, output_interface=output_interface,
                         output_dir=output_dir, data_selection=data_selection, custom_params=custom_params)
        return

    def check_data(self):

        return

    def custom_settings(self):
        for feature, filters in self.custom_params.items():
            # Plugin Duration in minutes
            if feature == 'plug_duration':
                # Add feature
                if 'plug_duration' not in self.df.columns:
                    self.df['plug_duration'] = (self.df['plug_out_datetime'] - self.df['plug_in_datetime']).dt.total_seconds() / 60

            # Energy supplied clipped between min and max values
            if feature == 'energy_supplied_clip':
                # Add feature
                if 'energy_supplied_clip' not in self.df.columns:
                    self.df['energy_supplied_clip'] = self.df['energy_supplied']
                # Apply filters
                self.df['energy_supplied_clip'] = np.clip(self.df['energy_supplied_clip'].values, a_min=filters['min'], a_max=filters['max'])

        return

    def run(self):
        """
        # Analysis description
        """
        output_dict = {}
        file_list = []

        if self.full_custom_mode == True and self.custom_params is None:  # Here we get general stats of the db (full_custom_mode=true and custom_params=None)
            # Run analysis
            table_stats = self.get_num_entries_tables()
            output_dict['table_stats'] = table_stats

            db_size = self.get_db_size()
            output_dict['db_size'] = db_size

            column_stats = self.get_column_statistics()
            output_dict['column_stats'] = column_stats

            # Show and save to file column stats
            if self.show_images or self.save_images:
                for table, col_stats_list in column_stats.items():
                    self.show_column_statistics(table_name=table, column_stats=col_stats_list)

            # Save stats to files
            if self.save_results:
                file_list.append(self.save_json_file(data=output_dict))
                file_list.append(self.save_xlsx_file(data=output_dict))
                file_list.append(self.save_pdf_file(data=output_dict, skip=['distribution']))

        elif self.custom_params['entity'] == 'dataset':
            print("Dataset statistics")

            # Filter dataset
            if self.custom_params['id']:
                dataset_name = self.input_interface.get_dataset_name_by_id(dataset_id=self.custom_params['id'])
                self.data_selection['datasets'] = [dataset_name]

            # Data gathering from input interface
            self.df = self.input_interface.get_data(data_selection=self.data_selection)
            # Data enrichment
            self.custom_settings()

            # Create report for the specific dataset
            for dataset_name in self.data_selection['datasets']:
                self.logger.info(f"{dataset_name} - Dataset statistics")
                filtered_df = self.df.loc[self.df['dataset_name'] == dataset_name]
                output_dict.update({dataset_name: {}})
                output_dict[dataset_name].update({'Dataset name': str(dataset_name)})
                output_dict[dataset_name].update({'Number of charges [#]': str(len(filtered_df))})
                output_dict[dataset_name].update({'Plug duration total [min]': str(round(filtered_df['plug_duration'].sum(), 2))})
                output_dict[dataset_name].update({'Plug duration min [min]': str(round(filtered_df['plug_duration'].min(), 2))})
                output_dict[dataset_name].update({'Plug duration max [min]': str(round(filtered_df['plug_duration'].max(), 2))})
                output_dict[dataset_name].update({'Plug duration mean [min]': str(round(filtered_df['plug_duration'].mean(), 2))})
                output_dict[dataset_name].update({'Energy supplied total [kWh]': str(round(filtered_df['energy_supplied_clip'].sum(), 2))})
                output_dict[dataset_name].update({'Energy supplied min [kWh]': str(round(filtered_df['energy_supplied_clip'].min(), 2))})
                output_dict[dataset_name].update({'Energy supplied max [kWh]': str(round(filtered_df['energy_supplied_clip'].max(), 2))})
                output_dict[dataset_name].update({'Energy supplied mean [kWh]': str(round(filtered_df['energy_supplied_clip'].mean(), 2))})

            # Create report for each single user
            for user_id in self.df['user_id'].unique():
                filtered_df = self.df.loc[self.df['user_id'] == user_id]
                output_dict.update({str(user_id): {}})
                output_dict[str(user_id)].update({'User id': str(user_id)})
                output_dict[str(user_id)].update({'Number of charges [#]': str(len(filtered_df))})
                output_dict[str(user_id)].update({'Plug duration total [min]': str(round(filtered_df['plug_duration'].sum(), 2))})
                output_dict[str(user_id)].update({'Plug duration min [min]': str(round(filtered_df['plug_duration'].min(), 2))})
                output_dict[str(user_id)].update({'Plug duration max [min]': str(round(filtered_df['plug_duration'].max(), 2))})
                output_dict[str(user_id)].update({'Plug duration mean [min]': str(round(filtered_df['plug_duration'].mean(), 2))})
                output_dict[str(user_id)].update({'Energy supplied total [kWh]': str(round(filtered_df['energy_supplied_clip'].sum(), 2))})
                output_dict[str(user_id)].update({'Energy supplied min [kWh]': str(round(filtered_df['energy_supplied_clip'].min(), 2))})
                output_dict[str(user_id)].update({'Energy supplied max [kWh]': str(round(filtered_df['energy_supplied_clip'].max(), 2))})
                output_dict[str(user_id)].update({'Energy supplied mean [kWh]': str(round(filtered_df['energy_supplied_clip'].mean(), 2))})
                output_dict[str(user_id)].update({'Average number of days between charges [#]': str(round(filtered_df['plug_in_datetime'].diff().dt.days.mean(), 2))})

            # Save stats to files
            if self.save_results:
                file_list.append(self.save_json_file(data=output_dict))
                file_list.append(self.save_pdf_file(data=output_dict))

        elif self.custom_params['entity'] == 'user':
            periods_in_days = [1, 7, 30, 180, 365, 365*50]
            now = datetime.now()
            for user_id in self.df['user_id'].unique():
                output_dict.update({str(user_id): {}})
                for period in periods_in_days:
                    output_dict[str(user_id)].update({str(period): {}})
                    start_period = now - timedelta(days=period)
                    filtered_df = self.df[(self.df['plug_in_datetime'] <= now) & (self.df['plug_in_datetime'] > start_period)]
                    number_of_charges = (filtered_df['user_id'] == user_id).sum()
                    total_plugs_duration = filtered_df[filtered_df['user_id'] == user_id]['plug_duration'].sum()
                    max_plug_duration = filtered_df[filtered_df['user_id'] == user_id]['plug_duration'].max()
                    min_plug_duration = filtered_df[filtered_df['user_id'] == user_id]['plug_duration'].min()
                    output_dict[str(user_id)][str(period)] = {
                        'number_of_charges': str(number_of_charges),
                        'total_plugs_duration': str(round(total_plugs_duration, 2) if number_of_charges > 0 else 0),
                        'max_plug_duration': str(round(max_plug_duration, 2) if number_of_charges > 0 else 0),
                        'min_plug_duration': str(round(min_plug_duration, 2) if number_of_charges > 0 else 0)
                    }

                # Save stats to files
                if self.save_results:
                    file_list.append(self.save_json_file(data=output_dict))

        elif self.custom_params['entity'] == 'chargingstation':
            # TODO
            print("Charging station statistics")
            # Save stats to files
            # if self.save_results:
            #     file_list.append(self.save_json_file(data=output_dict))

        self.logger.info(output_dict)

        # # TODO is this really needed?
        #  Encode file content into output_dict
        # if len(file_list) > 0:
        #     # Save all files to zip
        #     output_dict['files'] = {}
        #     buffer = self.create_zip_from_filepaths(filepaths=file_list)
        #     output_dict['files'].update(
        #         {'stats.zip': buffer.getvalue()})

        self.results = output_dict

        return self.results

    # Function to get database statistics
    def get_num_entries_tables(self):
        # Get cursor
        cur = self.input_interface.db.cursor()

        # Execute query to get the number of tables
        cur.execute("""
            SELECT schemaname, tablename
            FROM pg_tables
            WHERE schemaname IN ('evinsights');
        """)
        tables = cur.fetchall()

        # Calculate the number of rows for each table
        table_stats = {}
        for schema, table in tables:
            cur.execute(f'SELECT COUNT(*) FROM {schema}."{table}";')
            row_count = cur.fetchone()[0]
            table_stats.update({table: row_count})

        cur.close()

        return table_stats

    def get_db_size(self):
        # Get cursor
        cur = self.input_interface.db.cursor()

        # Execute query to get the size of the database
        cur.execute("""
            SELECT pg_database.datname, pg_size_pretty(pg_database_size(pg_database.datname))
            FROM pg_database
            WHERE pg_database.datname = current_database();
        """)

        db_size = cur.fetchone()
        db_size_dict = {db_size[0]: db_size[1]}

        cur.close()

        return db_size_dict

    def get_column_statistics(self):
        cur = self.input_interface.db.cursor()

        # Execute query to get the number of tables
        cur.execute("""
            SELECT schemaname, tablename
            FROM pg_tables
            WHERE schemaname IN ('evinsights');
        """)
        tables = cur.fetchall()

        table_column_stats = {}

        for schema, table in tables:
            # Execute query to get column names and data types
            cur.execute(f"""
                        SELECT column_name, data_type
                        FROM information_schema.columns
                        WHERE table_schema = '{schema}' AND table_name = '{table}';
                    """)
            columns = cur.fetchall()

            column_stats = []
            for column, data_type in columns:
                # Get the number of unique values
                cur.execute(f'SELECT COUNT(DISTINCT "{column}") FROM {schema}."{table}";')
                unique_count = cur.fetchone()[0]

                # Get the number of null values
                cur.execute(f'SELECT COUNT(*) FROM {schema}."{table}" WHERE "{column}" IS NULL;')
                null_count = cur.fetchone()[0]

                # Get the distribution of values (top 10 most frequent values)
                cur.execute(f"""
                            SELECT "{column}", COUNT(*)
                            FROM {schema}."{table}"
                            GROUP BY "{column}"
                            ORDER BY COUNT(*) DESC
                            LIMIT 10;
                        """)
                distribution = cur.fetchall()

                column_stats.append({
                    'table': table,
                    'column': column,
                    'data_type': data_type,
                    'unique_values': unique_count,
                    'null_values': null_count,
                    'distribution': distribution
                })
            table_column_stats[table] = column_stats

        cur.close()

        return table_column_stats

    def show_column_statistics(self, table_name, column_stats):
        # Create a DataFrame for unique and null values
        stats_df = pd.DataFrame(column_stats)

        # Bar chart for unique values and null values
        fig1 = go.Figure()
        fig1.add_trace(go.Bar(x=stats_df['column'], y=stats_df['unique_values'], name='Unique Values'))
        fig1.add_trace(go.Bar(x=stats_df['column'], y=stats_df['null_values'], name='Null Values'))

        fig1.update_layout(
            title=f'{table_name}: Unique and Null Values per Column',
            xaxis_title='Column',
            yaxis_title='Count',
            barmode='group'
        )

        if self.save_images:
            fig1.write_image(os.path.join(self.output_dir, f"{table_name}_uniques_and_nulls.png"), width=1080, height=720, scale=2)

        if self.show_images:
            fig1.show()

        # Pie charts for distribution of values (top 10 most frequent values) for each column
        for stat in column_stats:
            distribution = pd.DataFrame(stat['distribution'], columns=['value', 'count'])
            fig2 = px.pie(distribution, names='value', values='count', title=f"{table_name}: Value Distribution for {stat['column']}")

            if self.save_images:
                fig2.write_image(os.path.join(self.output_dir, f"{table_name}_{stat['column']}_distribution.png"),
                                 width=1080, height=720, scale=2)

            if self.show_images:
                fig2.show()

        return

    def create_zip_from_filepaths(self, filepaths: List[str]) -> BytesIO:
        """
        Create a ZIP archive in memory from a list of file paths.

        Args:
            filepaths: List of paths to the files to include in the ZIP archive.

        Returns:
            BytesIO: A BytesIO object containing the ZIP archive.
        """
        zip_buffer = BytesIO()
        with ZipFile(zip_buffer, 'w') as zipf:
            for path in filepaths:
                filename = os.path.basename(path)  # Use only the file name in the ZIP
                with open(path, 'rb') as f:
                    zipf.writestr(filename, f.read())
        zip_buffer.seek(0)
        return zip_buffer

    def save_json_file(self, data):
        """
        Save the output dictionary to a JSON file.
        """
        filename = os.path.join(self.output_dir, "stats.json")
        with open(filename, 'w') as file:
            pprint(data, stream=file)

        return filename

    def save_xlsx_file(self, data):
        """
        Save the output dictionary to a XLSX file.
        """
        filename = os.path.join(self.output_dir, "stats.xlsx")
        wb = Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Helper function to write a DataFrame to a sheet
        def write_df_to_sheet(workbook, sheet_name, dataframe):
            sheet = workbook.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(dataframe.itertuples(index=False, name=None), 2):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)
            for col_idx, col_name in enumerate(dataframe.columns, 1):
                sheet.cell(row=1, column=col_idx, value=col_name)

        # === TABLE_STATS SHEET ===
        table_stats = data["table_stats"]
        df_table_stats = pd.DataFrame([
            {"table_name": k, "records": v}
            for k, v in table_stats.items()
        ])
        write_df_to_sheet(wb, "table_stats", df_table_stats)

        # === COLUMN_STATS SHEET ===
        column_stats = data["column_stats"]
        records = []
        for dataset_name, columns in column_stats.items():
            for col in columns:
                record = {
                    "table": dataset_name,
                    "field": col["column"],
                    "type": col["data_type"],
                    "unique_values": col["unique_values"],
                    "null_values": col["null_values"],
                    "distribution_top_10": str(col["distribution"][:10])  # limit to 10 for readability
                }
                records.append(record)
        df_column_stats = pd.DataFrame(records)
        write_df_to_sheet(wb, "column_stats", df_column_stats)

        # === DB_SIZE SHEET ===
        db_size = data["db_size"]
        df_db_size = pd.DataFrame([{k: v} for k, v in db_size.items()])
        write_df_to_sheet(wb, "db_size", df_db_size)

        # Save the workbook to a file
        wb.save(filename)

        # Reopen the Excel file with openpyxl to adjust column widths
        workbook = load_workbook(filename)

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for column_cells in worksheet.columns:
                max_length = 0
                column_index = column_cells[0].column  # Numeric column index
                column_letter = get_column_letter(column_index)
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2  # Add padding
                worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save the final Excel file with adjusted column widths
        workbook.save(filename)
        return filename

    def save_pdf_file(self, data, skip=[]):
        """
        Save the output dictionary to a PDF file with tables in text format (generalized version).
        """
        filename = os.path.join(self.output_dir, "stats.pdf")

        def create_table(dataframe, col_widths=None, font_size=8):
            """
            Create a stylized text table from a DataFrame with word wrapping and auto row height.
            """
            styles = getSampleStyleSheet()

            # Style for normal cells
            styleN = ParagraphStyle(
                name="TableCell",
                parent=styles["Normal"],
                fontName="Helvetica",
                fontSize=font_size,
                leading=font_size + 2,
                textColor=colors.black
            )

            # Style for header
            styleHeader = ParagraphStyle(
                name="TableHeader",
                parent=styles["Normal"],
                fontName="Helvetica-Bold",
                fontSize=font_size,
                leading=font_size + 2,
                textColor=colors.white
            )

            # Convert headers and data to Paragraphs
            table_data = [[Paragraph(str(cell), styleHeader) for cell in dataframe.columns]]
            for _, row in dataframe.iterrows():
                table_data.append([Paragraph(str(cell), styleN) for cell in row])

            # Create table
            table = Table(table_data, colWidths=col_widths)

            # Define style
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ])

            # Alternating row backgrounds
            for i in range(1, len(table_data)):
                bg_color = colors.lightgrey if i % 2 == 1 else colors.white
                style.add('BACKGROUND', (0, i), (-1, i), bg_color)

            table.setStyle(style)
            return table

        # Create the PDF
        doc = SimpleDocTemplate(filename, pagesize=A4)
        content = []

        header_style = ParagraphStyle(
            name='HeaderStyle',
            fontSize=18,
            alignment=1,
            spaceAfter=10,
            fontName='Helvetica-Bold',
        )

        for key, value in data.items():
            # Normalize the data into a DataFrame
            if isinstance(value, dict):
                # Check if the values are lists
                if all(isinstance(v, list) for v in value.values()):
                    df = pd.DataFrame([
                        {**item} for k, v in value.items() for item in v if isinstance(item, dict)
                    ])
                    df.drop(columns=skip, inplace=True, errors='ignore')
                else:
                    df = pd.DataFrame([{'key': k, "value": v} for k, v in value.items()])
            elif isinstance(value, list):
                df = pd.DataFrame(value)
            else:
                continue  # skip unsupported types

            if df.empty:
                continue

            # Add section title
            content.append(Paragraph(f"{key.replace('_', ' ').title()}\n", style=header_style))

            # Set column widths based on number of columns
            col_count = len(df.columns)
            col_widths = [500 // col_count] * col_count  # Adjust width based on number of columns

            # Add table and page break
            content.append(create_table(df))
            content.append(PageBreak())

        doc.build(content)
        return filename

